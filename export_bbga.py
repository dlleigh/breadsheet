import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GRAMS_PER_OUNCE = 28.3495

def export_bbga_excel(
    formula_df: pd.DataFrame,
    calc,
    poolish_df: pd.DataFrame = None,
    sponge_df: pd.DataFrame = None,
    pre_fermented_flour_ratio: float = 0,
    output_path: str = 'formula_bbga.xlsx',
    title: str = ""
) -> None:
    """
    Export formula as BBGA-compliant Excel file with formulas and formatting

    Args:
        formula_df: Main recipe formula DataFrame
        calc: RecipeCalculator with batch info
        poolish_df: Optional poolish DataFrame with weights already calculated
        sponge_df: Optional sponge DataFrame with weights already calculated
        pre_fermented_flour_ratio: Fraction of total flour in preferment (0-1)
        output_path: Output file path
        title: Optional recipe title
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Formula"

    # Define colors (BBGA standard)
    DARK_BLUE = "00426A"
    ORANGE = "FF9900"
    YELLOW = "e2e28e"
    LIGHT_YELLOW = "FFF8DC"
    LIGHT_ORANGE = "FFD699"
    LIGHT_BLUE = "B8CCE4"

    # Define styles
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    light_yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    light_orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
    light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical='bottom')

    # No borders by default
    no_border = Border()

    # Vertical separators between sections
    separator_border = Border(right=Side(style='thin', color='000000'))

    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Determine if we have a preferment
    preferment_df = sponge_df if sponge_df is not None else poolish_df
    preferment_name = "SPONGE" if sponge_df is not None else "POOLISH" if poolish_df is not None else None

    # Convert grams to kilograms
    tdw_kg = calc.total_weight / 1000

    # Get flour percentage from formula
    flour_pct = 0
    # Import here to avoid circular dependency
    from breadsheet_utils import PrefermentCalculator
    if isinstance(calc, PrefermentCalculator):
        flour_pct = calc.get_flour_pct(formula_df)
    else:
        flour_pct = formula_df[formula_df.index.str.contains('flour', case=False)]['baker%'].sum()

    # Row 1: Title and TDW header
    current_row = 1
    if title:
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = title.upper()
        cell.font = Font(size=14, bold=True)
        cell.border = no_border
        current_row += 1

    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Total Dough Weight (TDW)"
    cell.fill = yellow_fill

    ws[f'C{current_row}'] = tdw_kg
    ws[f'C{current_row}'].fill = orange_fill
    ws[f'C{current_row}'].border = separator_border
    ws[f'C{current_row}'].number_format = '0.000" kg"'

    tdw_cell = f'C{current_row}'
    current_row += 1

    # Row 2: Preferment ratio if applicable
    if preferment_df is not None:
        #ws.merge_cells(f'D{current_row-1}:E{current_row-1}')
        cell = ws[f'D{current_row-1}']
        cell.value = "Total Flour Prefermented"
        cell.fill = yellow_fill
        cell.alignment = wrap_align

        ws[f'E{current_row-1}'] = pre_fermented_flour_ratio
        ws[f'E{current_row-1}'].fill = orange_fill
        ws[f'E{current_row-1}'].number_format = '0.00%'
        ws[f'E{current_row-1}'].border = separator_border

        ws[f'F{current_row-1}'].fill = yellow_fill
        ws[f'G{current_row-1}'].fill = yellow_fill
        preferment_ratio_cell = f'E{current_row-1}'

    # Section headers row (TOTAL FORMULA, SPONGE/POOLISH, FINAL DOUGH)
    section_headers_row = current_row

    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws[f'A{current_row}'] = "TOTAL FORMULA"
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = center_align
    ws[f'C{current_row}'].border = separator_border
    if preferment_df is not None:
        ws.merge_cells(f'D{current_row}:E{current_row}')
        ws[f'D{current_row}'] = preferment_name
        ws[f'D{current_row}'].fill = header_fill
        ws[f'D{current_row}'].font = header_font
        ws[f'D{current_row}'].alignment = center_align
        ws[f'E{current_row}'].border = separator_border
        ws.merge_cells(f'F{current_row}:G{current_row}')
        ws[f'F{current_row}'] = "FINAL DOUGH"
        ws[f'F{current_row}'].fill = header_fill
        ws[f'F{current_row}'].font = header_font
        ws[f'F{current_row}'].alignment = center_align
        ws[f'F{current_row}'].border = separator_border
    current_row += 1

    # Column headers row
    headers_row = current_row

    ws[f'A{current_row}'] = "Ingredients"
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].font = header_font
    ws[f'B{current_row}'] = "%"
    ws[f'B{current_row}'].fill = header_fill
    ws[f'B{current_row}'].font = header_font
    ws[f'B{current_row}'].alignment = center_align

    ws[f'C{current_row}'] = "kilograms"
    ws[f'C{current_row}'].fill = header_fill
    ws[f'C{current_row}'].font = header_font
    ws[f'C{current_row}'].border = separator_border
    ws[f'C{current_row}'].alignment = center_align

    if preferment_df is not None:
        # Preferment headers
        ws[f'D{current_row}'] = "%"
        ws[f'D{current_row}'].fill = header_fill
        ws[f'D{current_row}'].font = header_font
        ws[f'D{current_row}'].alignment = center_align

        ws[f'E{current_row}'] = "kilograms"
        ws[f'E{current_row}'].border = separator_border
        ws[f'E{current_row}'].fill = header_fill
        ws[f'E{current_row}'].font = header_font
        ws[f'E{current_row}'].alignment = center_align

        # Final Dough headers
        ws[f'F{current_row}'] = "Ingredients"
        ws[f'F{current_row}'].fill = header_fill
        ws[f'F{current_row}'].font = header_font
        ws[f'G{current_row}'] = "kilograms"
        ws[f'G{current_row}'].fill = header_fill
        ws[f'G{current_row}'].font = header_font
        ws[f'G{current_row}'].alignment = center_align

    current_row += 1
    formula_start_row = current_row

    # Get all unique ingredients from formula
    # If we have a preferment, formula_df already has it subtracted out in final dough
    # We need the original formula with all ingredients
    ingredients = list(formula_df.index)
    if preferment_df is not None and 'poolish' in ingredients:
        ingredients.remove('poolish')
    if preferment_df is not None and 'sponge' in ingredients:
        ingredients.remove('sponge')

    # Add rows for each ingredient in Total Formula
    formula_total_pct = formula_df['baker%'].sum()
    if preferment_df is not None:
        # Recalculate total without poolish/sponge row
        formula_total_pct = formula_df.drop(['poolish', 'sponge'], errors='ignore')['baker%'].sum()

    ingredient_rows = {}

    # Separate flour and non-flour ingredients
    flour_ingredients = [ing for ing in ingredients if 'flour' in ing.lower()]
    non_flour_ingredients = [ing for ing in ingredients if 'flour' not in ing.lower()]

    # Calculate where totals row will be
    # current_row + 1 (Total Flour row) + len(ingredients) + (1 if preferment for final dough row)
    totals_row_calc = formula_start_row + 1 + len(ingredients)  # +1 for Total Flour row at top
    if preferment_df is not None:
        totals_row_calc += 1  # Extra row for preferment in final dough

    # First add Total Flour row
    total_flour_row = current_row

    ws[f'A{current_row}'] = "Total Flour"
    ws[f'A{current_row}'].fill = light_blue_fill
    # Total flour percentage (will be calculated after individual flours are added)
    # Placeholder - will update with formula later
    ws[f'B{current_row}'].fill = light_blue_fill
    ws[f'B{current_row}'].number_format = '0.00%'

    # Total flour weight
    ws[f'C{current_row}'].fill = light_blue_fill
    ws[f'C{current_row}'].border = separator_border
    ws[f'C{current_row}'].number_format = '0.000" kg"'

    if preferment_df is not None:
        # Total flour in preferment
        ws[f'D{current_row}'].fill = light_blue_fill
        ws[f'D{current_row}'].number_format = '0.00%'

        ws[f'E{current_row}'].border = separator_border
        ws[f'E{current_row}'].fill = light_blue_fill
        ws[f'E{current_row}'].number_format = '0.000" kg"'

        # Total flour in final dough
        ws[f'F{current_row}'] = "Total Flour"
        ws[f'F{current_row}'].fill = light_blue_fill
        ws[f'G{current_row}'].fill = light_blue_fill
        ws[f'G{current_row}'].number_format = '0.000" kg"'

    current_row += 1

    # Now add flour ingredients
    row_index = 0
    for ingredient in flour_ingredients:
        # Alternate row colors for readability
        row_fill = yellow_fill if row_index % 2 == 0 else light_yellow_fill
        orange_row_fill = orange_fill if row_index % 2 == 0 else light_orange_fill

        # Total Formula columns
        ws[f'A{current_row}'] = ingredient.title()
        ws[f'A{current_row}'].fill = row_fill
            # Baker's percentage (editable orange cell)
        baker_pct = formula_df.loc[ingredient, 'baker%']
        ws[f'B{current_row}'] = baker_pct / 100
        ws[f'B{current_row}'].fill = orange_row_fill
        ws[f'B{current_row}'].number_format = '0.00%'

        # Kilograms (formula: baker% / total% * TDW)
        ws[f'C{current_row}'] = f'=B{current_row}/$B${totals_row_calc}*{tdw_cell}'
        ws[f'C{current_row}'].fill = row_fill
        ws[f'C{current_row}'].border = separator_border
        ws[f'C{current_row}'].number_format = '0.000" kg"'

        ingredient_rows[ingredient] = current_row

        # If we have a preferment, add preferment and final dough columns
        if preferment_df is not None:
            # Check if this ingredient is in the preferment
            if ingredient in preferment_df.index:
                # Preferment percentage (editable)
                pref_pct = preferment_df.loc[ingredient, 'baker%']
                ws[f'D{current_row}'] = pref_pct / 100
                ws[f'D{current_row}'].fill = orange_row_fill
                ws[f'D{current_row}'].number_format = '0.00%'

                # Preferment kilograms (formula)
                # This will be updated after we know the total row
                ws[f'E{current_row}'] = 0  # Placeholder
                ws[f'E{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
                ws[f'E{current_row}'].number_format = '0.000" kg"'
            else:
                # Blank cells in preferment columns
                ws[f'D{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
                ws[f'E{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
            # Final Dough ingredient name
            ws[f'F{current_row}'] = ingredient.title()
            ws[f'F{current_row}'].fill = row_fill
            # Final Dough kilograms (Total - Preferment)
            if ingredient in preferment_df.index:
                ws[f'G{current_row}'] = f'=C{current_row}-E{current_row}'
            else:
                ws[f'G{current_row}'] = f'=C{current_row}'
            ws[f'G{current_row}'].fill = row_fill
            ws[f'G{current_row}'].number_format = '0.000" kg"'

        current_row += 1
        row_index += 1

    # Update Total Flour row with formulas now that we know the flour row range
    flour_row_range = [ingredient_rows[ing] for ing in flour_ingredients]
    flour_start = min(flour_row_range)
    flour_end = max(flour_row_range)

    # Total flour percentage (sum of all flour percentages)
    ws[f'B{total_flour_row}'] = f'=SUM(B{flour_start}:B{flour_end})'

    # Total flour weight
    ws[f'C{total_flour_row}'] = f'=SUM(C{flour_start}:C{flour_end})'
    ws[f'C{total_flour_row}'].border = separator_border

    if preferment_df is not None:
        # Total flour in preferment
        flour_in_pref = [ing for ing in flour_ingredients if ing in preferment_df.index]
        if flour_in_pref:
            pref_flour_rows = [ingredient_rows[ing] for ing in flour_in_pref]
            pref_flour_start = min(pref_flour_rows)
            pref_flour_end = max(pref_flour_rows)

            ws[f'D{total_flour_row}'] = f'=SUM(D{pref_flour_start}:D{pref_flour_end})'
            ws[f'E{total_flour_row}'] = f'=SUM(E{pref_flour_start}:E{pref_flour_end})'
            ws[f'E{total_flour_row}'].border = separator_border

        # Total flour in final dough
        ws[f'G{total_flour_row}'] = f'=SUM(G{flour_start}:G{flour_end})'

    # Now add non-flour ingredients
    for ingredient in non_flour_ingredients:
        # Alternate row colors for readability (continue from flour ingredients)
        row_fill = yellow_fill if row_index % 2 == 0 else light_yellow_fill
        orange_row_fill = orange_fill if row_index % 2 == 0 else light_orange_fill

        # Total Formula columns
        ws[f'A{current_row}'] = ingredient.title()
        ws[f'A{current_row}'].fill = row_fill
            # Baker's percentage (editable orange cell)
        baker_pct = formula_df.loc[ingredient, 'baker%']
        ws[f'B{current_row}'] = baker_pct / 100
        ws[f'B{current_row}'].fill = orange_row_fill
        ws[f'B{current_row}'].number_format = '0.00%'

        # Kilograms (formula: baker% / total% * TDW)
        ws[f'C{current_row}'] = f'=B{current_row}/$B${totals_row_calc}*{tdw_cell}'
        ws[f'C{current_row}'].fill = row_fill
        ws[f'C{current_row}'].border = separator_border
        ws[f'C{current_row}'].number_format = '0.000" kg"'

        ingredient_rows[ingredient] = current_row

        # If we have a preferment, add preferment and final dough columns
        if preferment_df is not None:
            # Check if this ingredient is in the preferment
            if ingredient in preferment_df.index:
                # Preferment percentage (editable)
                pref_pct = preferment_df.loc[ingredient, 'baker%']
                ws[f'D{current_row}'] = pref_pct / 100
                ws[f'D{current_row}'].fill = orange_row_fill
                ws[f'D{current_row}'].number_format = '0.00%'

                # Preferment kilograms (formula)
                # This will be updated after we know the total row
                ws[f'E{current_row}'] = 0  # Placeholder
                ws[f'E{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
                ws[f'E{current_row}'].number_format = '0.000" kg"'
            else:
                # Blank cells in preferment columns
                ws[f'D{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
                ws[f'E{current_row}'].fill = row_fill
                ws[f'E{current_row}'].border = separator_border
            # Final Dough ingredient name
            ws[f'F{current_row}'] = ingredient.title()
            ws[f'F{current_row}'].fill = row_fill
            # Final Dough kilograms (Total - Preferment)
            if ingredient in preferment_df.index:
                ws[f'G{current_row}'] = f'=C{current_row}-E{current_row}'
            else:
                ws[f'G{current_row}'] = f'=C{current_row}'
            ws[f'G{current_row}'].fill = row_fill
            ws[f'G{current_row}'].number_format = '0.000" kg"'

        current_row += 1
        row_index += 1

    # Add preferment row in final dough if applicable
    if preferment_df is not None:
        # Continue alternating pattern
        row_fill = yellow_fill if row_index % 2 == 0 else light_yellow_fill

        # Fill empty cells in columns A-E with row color
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].fill = row_fill
        ws[f'C{current_row}'].fill = row_fill
        ws[f'C{current_row}'].border = separator_border
        for col in ['D']:
            ws[f'{col}{current_row}'].fill = row_fill
        ws[f'E{current_row}'].fill = row_fill
        ws[f'E{current_row}'].border = separator_border
        ws[f'F{current_row}'] = preferment_name.title()
        ws[f'F{current_row}'].fill = row_fill
        # Preferment total weight - sum of all preferment ingredients
        pref_start = formula_start_row
        pref_end = current_row - 1
        ws[f'G{current_row}'] = f'=SUM(E{pref_start}:E{pref_end})'
        ws[f'G{current_row}'].fill = row_fill
        ws[f'G{current_row}'].number_format = '0.000" kg"'

        current_row += 1

    # Totals row
    totals_row = current_row
    ws[f'A{totals_row}'] = "Totals"
    ws[f'A{totals_row}'].fill = header_fill
    ws[f'A{totals_row}'].font = header_font
    ws[f'A{totals_row}']
    # Total baker's percentage
    ws[f'B{totals_row}'] = f'=SUM(B{formula_start_row}:B{totals_row-1})'
    ws[f'B{totals_row}'].fill = header_fill
    ws[f'B{totals_row}'].font = header_font
    ws[f'B{totals_row}'].number_format = '0.00%'

    # Total weight (should equal TDW)
    ws[f'C{totals_row}'] = f'=SUM(C{formula_start_row}:C{totals_row-1})'
    ws[f'C{totals_row}'].fill = header_fill
    ws[f'C{totals_row}'].font = header_font
    ws[f'C{totals_row}'].border = separator_border
    ws[f'C{totals_row}'].number_format = '0.000" kg"'

    if preferment_df is not None:
        # Preferment totals
        ws[f'D{totals_row}'] = f'=SUM(D{formula_start_row}:D{totals_row-1})'
        ws[f'D{totals_row}'].fill = header_fill
        ws[f'D{totals_row}'].font = header_font
        ws[f'D{totals_row}'].number_format = '0.00%'

        ws[f'E{totals_row}'] = f'=SUM(E{formula_start_row}:E{totals_row-1})'
        ws[f'E{totals_row}'].fill = header_fill
        ws[f'E{totals_row}'].font = header_font
        ws[f'E{totals_row}'].border = separator_border
        ws[f'E{totals_row}'].number_format = '0.000" kg"'

        # Final dough label
        ws[f'F{totals_row}'] = ""
        ws[f'F{totals_row}'].fill = header_fill
        ws[f'F{totals_row}'].font = header_font
        # Final dough total
        ws[f'G{totals_row}'] = f'=SUM(G{formula_start_row}:G{totals_row-1})'
        ws[f'G{totals_row}'].fill = header_fill
        ws[f'G{totals_row}'].font = header_font
        ws[f'G{totals_row}'].number_format = '0.000" kg"'

        # Now update preferment formulas
        # Get flour rows
        flour_rows = [ingredient_rows[ing] for ing in ingredient_rows if 'flour' in ing.lower()]

        # Calculate total flour percentage in preferment (sum of flour rows in column D)
        flour_pct_sum = '+'.join([f'D{r}' for r in flour_rows if list(ingredient_rows.keys())[list(ingredient_rows.values()).index(r)] in preferment_df.index])

        # Total flour weight in formula
        total_flour_sum = '+'.join([f'C{r}' for r in flour_rows])

        # For each preferment ingredient, create proper formula
        for ingredient in preferment_df.index:
            if ingredient in ingredient_rows:
                row = ingredient_rows[ingredient]
                if 'flour' in ingredient.lower():
                    # Flour in preferment: (this_flour% / total_flour%) * total_flour_kg * preferment_ratio
                    ws[f'E{row}'] = f'=D{row}/({flour_pct_sum})*({total_flour_sum})*{preferment_ratio_cell}'
                else:
                    # Non-flour: (ingredient% / 100%) * total_preferment_flour_kg
                    # total_preferment_flour_kg = sum of flour rows in column E
                    pref_flour_sum = '+'.join([f'E{r}' for r in flour_rows if list(ingredient_rows.keys())[list(ingredient_rows.values()).index(r)] in preferment_df.index])
                    ws[f'E{row}'] = f'=D{row}/1*({pref_flour_sum})'

                ws[f'E{row}'].number_format = '0.000" kg"'

    # Set row & column widths
    ws.row_dimensions[2].height = 35  # TDW row
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    if preferment_df is not None:
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12

    # Save workbook
    wb.save(output_path)
    print(f"BBGA formula exported to {output_path}")
